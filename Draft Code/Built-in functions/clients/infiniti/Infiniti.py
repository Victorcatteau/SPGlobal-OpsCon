#!/usr/bin/env python
# -*- coding: utf-8 -*- 

"""--------------------------------------------------------------------
INFINITI WARRANTY PROJECT


Hugo PERRIN
------------------------------------------------------------------------
"""


#===========================================================================================
# NEXT STEPS : PUT THE HTML MAIL IN TEXT FILES WITH {1} {2} ... TO HAVE A CLEARER SCRIPT
#===========================================================================================


#===========================================================================================
#====================================== LIBRARIES ==========================================



import pandas as pd
from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
import numpy as np
from email import *
import win32com.client as win32
from datetime import date, datetime
import smtplib,ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText           
from email.utils import COMMASPACE, formatdate
from email import encoders
import email.mime.application
import os




#===========================================================================================
#====================================== FUNCTIONS ==========================================



def style_range(worksheet, row_min, row_max, col_min, col_max, border=None, fill=None, font=None, alignment=None):
    for i in np.arange(row_min, row_max+1):
        for j in np.arange(col_min, col_max+1):
            worksheet.cell(row=i, column=j).fill = fill
            worksheet.cell(row=i, column=j).font = font
            worksheet.cell(row=i, column=j).border = border
            worksheet.cell(row=i, column=j).alignment = alignment




def no_quadrillage(worksheet, row_min, row_max, col_min, col_max):
    style_range(worksheet, row_min = 1, row_max = 200,col_min = 1, col_max = 100,
                fill = PatternFill(fill_type='none', start_color='77B5FE'),
                font = Font(name='Calibri', size=10, bold=False),
                border = Border(left=Side(border_style="thin", color='FFFFFF'),
                            right=Side(border_style="thin", color='FFFFFF'),
                            top=Side(border_style="thin", color='FFFFFF'),
                            bottom=Side(border_style="thin", color='FFFFFF')),
                alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=False))




def set_columns_width(worksheet, width, columns):
        i = 0
        for col in worksheet.columns:
            if i in columns:
                column = col[0].column 
                for cell in col:
                    worksheet.column_dimensions[column].width = width
            i += 1





def Outlook_email_sender(receiver_mail, subject, text, attachments = None):
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)

    # Receiver
    mail.To = receiver_mail

    # Subject
    mail.Subject = subject

    # Text 
    mail.HTMLBody = text

    # Attachment
    if attachments is not None:
        for attachment in attachments:
            mail.Attachments.Add(attachment)

    # Sending
    mail.Send()





def send_mail(send_from,
              send_to,
              subject,
              text,
              files,
              names,
              server,
              port,
              username='',
              password='',
              cc = None,
              isTls=True):

        msg = MIMEMultipart()
        msg['From'] = send_from
        msg['To'] = send_to
        msg['Date'] = formatdate(localtime = True)
        msg['Subject'] = subject
        msg.attach(MIMEText(text, "html"))

        if cc is not None:
            msg['Cc'] = cc
            
        for name,f in zip(names,files):
            fp=open(f,'rb')
            att = email.mime.application.MIMEApplication(fp.read())
            fp.close()
            att.add_header('Content-Disposition','attachment',filename=name)
            msg.attach(att)
            
        smtp = smtplib.SMTP(server, port)
        if isTls:
            smtp.starttls()
        smtp.login(username,password)
        if cc is not None:
            smtp.sendmail(send_from, [y for x in [send_to.split(";"), [cc]] for y in x], msg.as_string())
        else:
            smtp.sendmail(send_from, send_to.split(";"), msg.as_string())

        smtp.quit()
        
        print(">> Email sent to {}".format(send_to.split(";")))




#===========================================================================================
#==================================== DICTIONARIES =========================================



french_dictionnary = {"January" : "Janvier",
                    "February" : "Février",
                    "March" : "Mars",
                    "April" : "Avril",
                    "May" : "Mai",
                    "June" : "Juin",
                    "July" : "Muillet",
                    "August" : "Août",
                    "September" : "Septembre",
                    "October" : "Octobre",
                    "November" : "Novembre",
                    "December" : "Décembre"}


italian_dictionnary = {"January" : "Gennaio",
                    "February" : "Febbraio",
                    "March" : "Marzo",
                    "April" : "Aprile",
                    "May" : "Maggio",
                    "June" : "Giugno",
                    "July" : "Luglio",
                    "August" : "Agosto",
                    "September" : "Settembre",
                    "October" : "Ottobre",
                    "November" : "Novembre",
                    "December" : "Dicembre"}



#------------------------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------------
#-------------------------------------------- MAIN CLASS ----------------------------------------------------
#------------------------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------------



class Dealer(object):
    """
    This class aims at automating the monthly treatment of Inifniti Warranty team data 

    """

    def __init__(self, name, main, attachments_path, contact):


        self.name = name
        self.get_country(contact) # launch also self.get_month()
        self.get_center(contact)
        self.global_registration = main
        self.get_attachment(attachments_path)



    #-----------------------------------------------------------------------------------------------------
    #-- GETTERS



    def get_month(self):
        date = datetime.now()

        # Minuscule

        if self.country == "UK":
            self.month = date.strftime("%B").lower()
        if self.country == "France":
            self.month = french_dictionnary[date.strftime("%B")].lower()
        if self.country == "Italy":
            self.month = italian_dictionnary[date.strftime("%B")].lower()

        # Majuscule

        if self.country == "UK":
            self.maj_month = date.strftime("%B")
        if self.country == "France":
            self.maj_month = french_dictionnary[date.strftime("%B")]
        if self.country == "Italy":
            self.maj_month = italian_dictionnary[date.strftime("%B")]



    def get_country(self, contact_database):
        self.country = contact_database["Country"][contact_database["Dealer"] == self.name].values[0]
        self.get_month()



    def get_compliance_ratio(self):
        if self.eligible_number == 0:
            self.compliance_ratio = 0
        else:
            self.compliance_ratio = round(self.WC_number/self.eligible_number*100,1)



    def get_RNSvsEligible(self, national_base, country_VIN_name):
        self.eligible = [vin for vin in list(self.global_registration["Vin"][self.global_registration["Dealer"] == self.name]) if vin in list(national_base[country_VIN_name])]
        self.eligible_number = len(self.eligible)
        self.RNS_number = len([vin for vin in list(self.global_registration["Vin"][self.global_registration["Dealer"] == self.name]) if vin not in list(national_base[country_VIN_name])])



    def get_lacking_WSD(self):
        self.lacking_WSD = [vin for vin in self.eligible if vin in list(self.global_registration["Vin"][self.global_registration["NEWS Data - WSD"].isnull()])]
        self.WSD_number = self.eligible_number - len(self.lacking_WSD)



    def get_lacking_WC(self):
        self.lacking_WC = [vin for vin in self.eligible if vin in list(self.global_registration["Vin"][self.global_registration['Maindata_PSWR_CertificatePrinted.VIN'] == "N"])]
        self.WC_number = self.eligible_number - len(self.lacking_WC)



    def get_contact_mail(self, contact_database):
        if pd.isnull(contact_database["Mail"][contact_database["Dealer"] == self.name].values[0]) is True:
            self.contact = None
        else:
            self.contact = contact_database["Mail"][contact_database["Dealer"] == self.name].values[0]



    def get_copy(self, contact_database):
        if pd.isnull(contact_database["Copie"][contact_database["Dealer"] == self.name].values[0]) is True:
            self.copy = None
        else:
            self.copy = contact_database["Copie"][contact_database["Dealer"] == self.name].values[0]



    def get_center(self, contact_database):
        self.center = contact_database["Center"][contact_database["Dealer"] == self.name].values[0]



    def get_attachment(self, attachments_path):

        if self.country == "UK":
            self.attachment_path = attachments_path + "\\Warranty report - {} - {}.xlsx".format(self.center, self.maj_month)
            self.attachment_name = "Warranty report - {} - {}.xlsx".format(self.center, self.maj_month)
            self.object = "Warranty report - {} - {}.xlsx".format(self.center, self.maj_month)
        if self.country == "France":
            self.attachment_path = attachments_path + "\\Rapport de garantie - {} - {}.xlsx".format(self.center, self.maj_month)
            self.attachment_name = "Rapport de garantie - {} - {}.xlsx".format(self.center, self.maj_month)
            self.object = "Rapport de garantie - {} - {}.xlsx".format(self.center, self.maj_month)
        if self.country == "Italy":
            self.attachment_path = attachments_path + "\\Rapporto di garanzia - {} - {}.xlsx".format(self.center, self.maj_month)
            self.attachment_name = "Rapporto di garanzia - {} - {}.xlsx".format(self.center, self.maj_month)
            self.object = "Rapporto di garanzia - {} - {}.xlsx".format(self.center, self.maj_month)





    #-----------------------------------------------------------------------------------------------------
    #-- CREATING AND SAVING DATA



    def create_attachment(self):
        self.workbook = Workbook()
        ##--------------------------------------------------------------------------------------------
        ## FIRST SHEET 
        self.workbook.active.title = "Warranty Start Date"
        no_quadrillage(self.workbook["Warranty Start Date"], row_min = 1, row_max = len(self.lacking_WSD) + 100, col_min = 1, col_max = 50)
        set_columns_width(self.workbook["Warranty Start Date"], 30, [1])
        set_columns_width(self.workbook["Warranty Start Date"], 5, [0])

        if len(self.lacking_WSD) != 0:

            style_range(self.workbook["Warranty Start Date"], row_min = 2, row_max = 2, col_min = 2, col_max = 2,
                fill = PatternFill(fill_type='solid', start_color = "FFFFFF"),
                font = Font(name='Calibri', size=10, bold=True, color = 'B9121B'),
                border = Border(left=Side(border_style= 'none'),
                                right=Side(border_style="none"),
                                top=Side(border_style="none"),
                                bottom=Side(border_style="none")),
                alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=False))

            style_range(self.workbook["Warranty Start Date"], row_min = 3, row_max = 4, col_min = 2, col_max = 2,
                fill = PatternFill(fill_type='solid', start_color = "FFFFFF"),
                font = Font(name='Calibri', size=10, bold=False, color = 'B9121B'),
                border = Border(left=Side(border_style= 'none'),
                                right=Side(border_style="none"),
                                top=Side(border_style="none"),
                                bottom=Side(border_style="none")),
                alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=False))

            for i in range(len(self.lacking_WSD)):
                self.workbook["Warranty Start Date"].cell(row = i+7, column = 2, value = self.lacking_WSD[i])

            style_range(self.workbook["Warranty Start Date"], row_min = 6, row_max = 6, col_min = 2, col_max = 2,
                fill = PatternFill(fill_type='solid', start_color='000000'),
                font = Font(name='Calibri', size=10, bold=False, color = 'FEFEFE'),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                            right=Side(border_style="thin", color='FF000000'),
                            top=Side(border_style="thin", color='FF000000'),
                            bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

            style_range(self.workbook["Warranty Start Date"], row_min = 7, row_max = 6+len(self.lacking_WSD), col_min = 2, col_max = 2,
                fill = PatternFill(fill_type='none'),
                font = Font(name='Calibri', size=10, bold=False),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                            right=Side(border_style="thin", color='FF000000'),
                            top=Side(border_style="thin", color='FF000000'),
                            bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

            if self.country == "UK":
                self.workbook["Warranty Start Date"].cell(row = 2, column = 2, value = "Actions required :")
                self.workbook["Warranty Start Date"].cell(row = 3, column = 2, value = "1 - Input the Warranty Start Date into the Infiniti Warranty system. Please note that the Warranty Start Date is the date of customer handover or the date on which the vehicle is put into use, whichever comes first.")
                self.workbook["Warranty Start Date"].cell(row = 4, column = 2, value = "2 - Print the Warranty certificate. Then provide the warranty certificate to the customer next time he/she visits your Infiniti centre.")
                self.workbook["Warranty Start Date"].cell(row = 6, column = 2, value = "VINs of Vehicles without Warranty Starting Date and Warranty Certificate")
                self.workbook["Warranty Start Date"].title = "Warranty start date"

            elif self.country == "France":
                self.workbook["Warranty Start Date"].cell(row = 2, column = 2, value = "Actions requises :")
                self.workbook["Warranty Start Date"].cell(row = 3, column = 2, value = "1 - Saisissez la date de début de garantie dans votre système de garantie Infiniti. Il est à noter que la date de début de garantie est la date de remise du véhicule au client ou la date à laquelle le véhicule est mis en circulation.")
                self.workbook["Warranty Start Date"].cell(row = 4, column = 2, value = "2 - Imprimez le certificat de garantie et remettez-le au client lors de sa prochaine visite au Centre Infiniti.")
                self.workbook["Warranty Start Date"].cell(row = 6, column = 2, value = "VINs des véhicules dont la date de garantie et le certificat de garantie sont manquants")
                self.workbook["Warranty Start Date"].title = "Date de début de garantie"

            elif self.country == "Italy":
                self.workbook["Warranty Start Date"].cell(row = 2, column = 2, value = "Interventi richiesti :")
                self.workbook["Warranty Start Date"].cell(row = 3, column = 2, value = "1 - Inserire la data di inizio della garanzia nel sistema garanzie Infiniti. Si prega di notare che la data di inizio della garanzia coincide con la data di consegna al cliente o con la data di primo utilizzo del veicolo, a seconda di quale condizione si verifica per prima.")
                self.workbook["Warranty Start Date"].cell(row = 4, column = 2, value = "2 - Stampare il certificato di garanzia. Fornire quindi il certificato di garanzia al cliente alla prima occasione in cui questi fa visita al vostro Centro Infiniti.")
                self.workbook["Warranty Start Date"].cell(row = 6, column = 2, value = "VINs senza data di inizio della garanzia e certificato di garanzia stampato")
                self.workbook["Warranty Start Date"].title = "Data di inizio della garanzia"

        else: 

            style_range(self.workbook["Warranty Start Date"], row_min = 2, row_max = 2, col_min = 2, col_max = 2,
                fill = PatternFill(fill_type='solid', start_color='000000'),
                font = Font(name='Calibri', size=10, bold=False, color = 'FEFEFE'),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                            right=Side(border_style="thin", color='FF000000'),
                            top=Side(border_style="thin", color='FF000000'),
                            bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

            if self.country == "UK":
                self.workbook["Warranty Start Date"].cell(row = 2, column = 2, value = "All the vehicles sold have a Warranty Start Date")
                self.workbook["Warranty Start Date"].title = "Warranty start date"
            elif self.country == "France":
                self.workbook["Warranty Start Date"].cell(row = 2, column = 2, value = "Tous les véhicules vendues ont une date de début de garantie")
                self.workbook["Warranty Start Date"].title = "Date de début de garantie"
            elif self.country == "Italy":
                self.workbook["Warranty Start Date"].cell(row = 2, column = 2, value = "Tutti i veicoli venduti hanno data di inizio della garanzia")
                self.workbook["Warranty Start Date"].title = "Data di inizio della garanzia"



        ##--------------------------------------------------------------------------------------------
        ## SECOND SHEET 
        self.workbook.create_sheet("Warranty Certificate")
        no_quadrillage(self.workbook["Warranty Certificate"], row_min = 1, row_max = len(self.lacking_WC) + 100, col_min = 1, col_max = 50)


        style_range(self.workbook["Warranty Certificate"], row_min = 2, row_max = 2, col_min = 2, col_max = 2,
                fill = PatternFill(fill_type='solid', start_color = "FFFFFF"),
                font = Font(name='Calibri', size=10, bold=True, color = 'B9121B'),
                border = Border(left=Side(border_style= 'none'),
                                right=Side(border_style="none"),
                                top=Side(border_style="none"),
                                bottom=Side(border_style="none")),
                alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=False))

        style_range(self.workbook["Warranty Certificate"], row_min = 3, row_max = 3, col_min = 2, col_max = 2,
                fill = PatternFill(fill_type='solid', start_color = "FFFFFF"),
                font = Font(name='Calibri', size=10, bold=False, color = 'B9121B'),
                border = Border(left=Side(border_style= 'none'),
                                right=Side(border_style="none"),
                                top=Side(border_style="none"),
                                bottom=Side(border_style="none")),
                alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=False))


        for i in range(len(self.lacking_WC)):
            self.workbook["Warranty Certificate"].cell(row = i+6, column = 2, value = self.lacking_WC[i])


        style_range(self.workbook["Warranty Certificate"], row_min = 5, row_max = 5, col_min = 2, col_max = 2,
            fill = PatternFill(fill_type='solid', start_color='000000'),
            font = Font(name='Calibri', size=10, bold=False, color = 'FEFEFE'),
            border = Border(left=Side(border_style="thin", color='FF000000'),
                            right=Side(border_style="thin", color='FF000000'),
                            top=Side(border_style="thin", color='FF000000'),
                            bottom=Side(border_style="thin", color='FF000000')),
            alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

        style_range(self.workbook["Warranty Certificate"], row_min = 6, row_max = 5+len(self.lacking_WC), col_min = 2, col_max = 2,
            fill = PatternFill(fill_type='none'),
            font = Font(name='Calibri', size=10, bold=False),
            border = Border(left=Side(border_style="thin", color='FF000000'),
                            right=Side(border_style="thin", color='FF000000'),
                            top=Side(border_style="thin", color='FF000000'),
                            bottom=Side(border_style="thin", color='FF000000')),
            alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

        set_columns_width(self.workbook["Warranty Certificate"], 30, [1])
        set_columns_width(self.workbook["Warranty Certificate"], 5, [0])

        if self.country == "UK":
            self.workbook["Warranty Certificate"].cell(row = 2, column = 2, value = "Actions required :")
            self.workbook["Warranty Certificate"].cell(row = 3, column = 2, value = "Print the Warranty certificate. Then provide the warranty certificate to the customer next time he/she visits your Infiniti centre.")
            self.workbook["Warranty Certificate"].cell(row = 5, column = 2, value = "VINs of Vehicles without Warranty Certificate")
            self.workbook["Warranty Certificate"].title = "Warranty certificate"

        elif self.country == "France":
            self.workbook["Warranty Certificate"].cell(row = 2, column = 2, value = "Actions requises :")
            self.workbook["Warranty Certificate"].cell(row = 3, column = 2, value = "Imprimez le certificat de garantie et remettez-le au client lors de sa prochaine visite au Centre Infiniti.")
            self.workbook["Warranty Certificate"].cell(row = 5, column = 2, value = "VINs des véhicules dont le certificat de garantie n'a pas été imprimé")
            self.workbook["Warranty Certificate"].title = "Certificat de garantie"

        elif self.country == "Italy":
            self.workbook["Warranty Certificate"].cell(row = 2, column = 2, value = "Interventi richiesti :")
            self.workbook["Warranty Certificate"].cell(row = 3, column = 2, value = "Stampare il certificato di garanzia. Fornire quindi il certificato di garanzia al cliente alla prima occasione in cui questi fa visita al vostro Centro Infiniti.")
            self.workbook["Warranty Certificate"].cell(row = 5, column = 2, value = "VINs senza certificato di garanzia stampato")
            self.workbook["Warranty Certificate"].title = "Certificato di garanzia"



    def save_attachment(self):
        self.workbook.save(filename = self.attachment_path)



    #-----------------------------------------------------------------------------------------------------
    #-- SENDING DATA



    def send_email(self,contact_database):
        self.get_text(contact_database)
        self.get_contact_mail(contact_database)
        self.get_copy(contact_database)

        if self.country != "Italy":
            if self.contact is not None:
                if self.compliance_ratio != 100:
                    "Envoyer mail général avec le fichier joint"
                    
                    if self.country == "France":
                        self.bulletin_path = "S:/121. Infiniti/PDF/Exigences de garantie en pré-vente (PSWR).pdf"
                        self.bulletin_name = "Exigences de garantie en pré-vente (PSWR).pdf"
                    elif self.country == "UK":
                        self.bulletin_path = "S:/121. Infiniti/PDF/Pre-Sales Warranty Requirements (PSWR).pdf"
                        self.bulletin_name = "Pre-Sales Warranty Requirements (PSWR).pdf"
                    elif self.country == "Italy":
                        self.bulletin_path = "S:/121. Infiniti/PDF/Modulo Requisiti garanzia prevendita (PSWR).pdf"
                        self.bulletin_name = "Modulo Requisiti garanzia prevendita (PSWR).pdf"

                    
                    send_mail(send_from = 'Infiniti@ekimetrics.com',
                            send_to= self.contact,
                            subject = self.object,
                            text = self.text,
                            files = [self.attachment_path,self.bulletin_path],
                            names = [self.attachment_name,self.bulletin_name],
                            server = 'ssl0.ovh.net',
                            port = 25,
                            username ='Infiniti@ekimetrics.com',
                            password='EKLvfbcINF136',
                            cc = self.copy)

                    print(">> For {}".format(self.center))

                    #Outlook_email_sender(self.contact, "Warranty Report : {}".format(self.month), self.text, attachments = [self.attachment_path,self.bulletin_path])
            
                else:
                    "envoyer mail 100%"

                    send_mail(send_from = 'Infiniti@ekimetrics.com',
                            send_to= self.contact,
                            subject = self.object,
                            text = self.text,
                            files = [],
                            names = [],
                            server = 'ssl0.ovh.net',
                            port = 25,
                            username ='Infiniti@ekimetrics.com',
                            password='EKLvfbcINF136',
                            cc = self.copy)

                    print(">> For {}".format(self.center))

                    #Outlook_email_sender(self.contact, "Warranty Report : {}".format(self.month), self.text)



    def get_text(self, contact_database):

        if self.country == "UK":

            if self.compliance_ratio == 100:

    ### TEXTE ANGLAIS 100%

                self.text =  """<html>

    <head>

    </head>

    <body lang=FR style='tab-interval:36.0pt'>

    <div class=WordSection1>

    <p class=MsoNormal><span lang=EN-US>Dear Infiniti Center """ + self.center + """</span></p>

    <p class=MsoNormal><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>We are at the
    end of the month of """+ self.month + """ and we would like to share with you an update of your
    <b style='mso-bidi-font-weight:normal'>Compliance ratio* for Warranty
    certificate printed</b> for all eligible vehicles.</span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>As you know <b
    style='mso-bidi-font-weight:normal'><u>the objective of Compliance ratio is
    100%.<o:p></o:p></u></b></span></p>

    <p class=MsoNormal align=center style='text-align:center'><b style='mso-bidi-font-weight:
    normal'><span lang=EN-US><o:p>&nbsp;</o:p></span></b></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>At
    the end of the month your Compliance ratio is : <b style='mso-bidi-font-weight:normal'> 100% <o:p></o:p></u></b></span></p>

    <p class=MsoNormal><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>Thank you for adhering to the Pre-Sales Warranty Requirements 
    for vehicles sold / handed over to the customer. You achieve the objective for this month. Continue your effort, to make sure 
    you will keep this high quality of service.</span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>Best regards,</span></p>

    <p class=MsoNormal style='text-align:justify'><b style='mso-bidi-font-weight:
    normal'>Infiniti Europe Aftersales team<o:p></o:p></b></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-GB
    style='font-size:8.0pt;mso-bidi-font-size:11.0pt;line-height:107%;mso-ansi-language:
    EN-GB'>*Compliance ratio = Amount of eligible vehicles with Printed Warranty
    certificate in system / The Total of eligible vehicles<o:p></o:p></span></p>

    </div>

    </body>

    </html>"""

    ### TEXTE ANGLAIS NOT 100%

            else:
                self.text = """<html>

    <head>

    </head>

    <body lang=EN-US link="#0563C1" vlink="#954F72" style='tab-interval:36.0pt'>

    <div class=WordSection1>

    <p class=MsoNormal>Dear Infiniti Center """ + self.center + """,</p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'>As you know all vehicles sold /
    handed over to customers must have the Warranty <span class=GramE>Pre Sales
    Requirements</span> fully completed and the Warranty certificate should be
    printed and attached to the warranty booklet. The warranty certificate cannot
    be printed without completing the Pre-Sales Warranty requirements
    (WB-IEU-17-002)</p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'>Your current <b style='mso-bidi-font-weight:
    normal'>Compliance ratio* for Warranty certificate printed</b> for all eligible
    vehicles is:<b style='mso-bidi-font-weight:normal'> """ + str(self.compliance_ratio) + """%</b></p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'><b style='mso-bidi-font-weight:
    normal'><u>The objective for Compliance ratio is 100%.<o:p></o:p></u></b></p>

    <p class=MsoNormal style='text-align:justify'>Attached you will find a list of
    VINs which have already been handed over to the customer but they still do not
    have a Warranty Date in the system and/or a Warranty certificate printed. We
    have separated the report in two sections.</p>

    <p class=MsoListParagraphCxSpFirst style='margin-left:38.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-18.0pt;mso-list:l1 level1 lfo5'><![if !supportLists]><span
    style='font-family:Symbol;mso-fareast-font-family:Symbol;mso-bidi-font-family:
    Symbol'><span style='mso-list:Ignore'>·<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>Section 1 : VINs which have their <b
    style='mso-bidi-font-weight:normal'>Warranty Start Date </b>missing and do NOT
    have the <b style='mso-bidi-font-weight:normal'>Warranty certificate printed</b></p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:74.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-18.0pt;mso-list:l1 level2 lfo5'><![if !supportLists]><span
    style='font-family:"Courier New";mso-fareast-font-family:"Courier New"'><span
    style='mso-list:Ignore'>o<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;
    </span></span></span><![endif]><b style='mso-bidi-font-weight:normal'>Actions
    required : </b></p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:110.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-30.0pt;mso-list:l1 level3 lfo5'><![if !supportLists]><span
    style='mso-bidi-font-family:Calibri;mso-bidi-theme-font:minor-latin'><span
    style='mso-list:Ignore'>1.<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>Input the Warranty Start Date into the Infiniti
    Warranty system. Please note that the Warranty Start Date is the date of
    customer handover or the date on which the vehicle is put into use, whichever
    comes first.</p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:110.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-30.0pt;line-height:105%;mso-list:l1 level3 lfo5'><![if !supportLists]><span
    style='mso-bidi-font-family:Calibri;mso-bidi-theme-font:minor-latin'><span
    style='mso-list:Ignore'>2.<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>Print the Warranty certificate. Then provide the
    warranty certificate to the customer next time he/she visits your Infiniti
    centre.</p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:38.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-18.0pt;mso-list:l1 level1 lfo5'><![if !supportLists]><span
    style='font-family:Symbol;mso-fareast-font-family:Symbol;mso-bidi-font-family:
    Symbol'><span style='mso-list:Ignore'>·<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>Section 2 : VINs which need to have their<b
    style='mso-bidi-font-weight:normal'> Warranty certificate</b> printed</p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:74.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-18.0pt;mso-list:l1 level2 lfo5'><![if !supportLists]><span
    style='font-family:"Courier New";mso-fareast-font-family:"Courier New"'><span
    style='mso-list:Ignore'>o<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;
    </span></span></span><![endif]><b style='mso-bidi-font-weight:normal'>Actions
    required :<o:p></o:p></b></p>

    <p class=MsoListParagraphCxSpLast style='margin-left:110.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-30.0pt;line-height:105%;mso-list:l1 level3 lfo5'><![if !supportLists]><span
    style='mso-bidi-font-family:Calibri;mso-bidi-theme-font:minor-latin'><span
    style='mso-list:Ignore'>1.<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>Print the Warranty certificate. Then provide the
    warranty certificate to the customer next time he/she visits your Infiniti
    centre.</p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'>Attached you will find the Pre-Sales
    Warranty Requirements (PSWR) with all the details of mandatory tasks that must
    be completed before a vehicle is handed over to the final Customer, in this way
    the Warranty coverage can be validated and Customer’s new vehicle is covered by
    Roadside Assistance. Please ensure that all warranty requirements are completed
    <b style='mso-bidi-font-weight:normal'><u>prior</u></b> to customer handover.</p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=3DMsoNormal><span class=3DGramE><span lang=3DEN-US style=3D'mso-an=
    si-language:
    EN-US'>If</span></span><span lang=3DEN-US style=3D'mso-ansi-language:EN-US'=
    > you
    have any questions please email <a href="mailto:Infiniti@ekimetrics.com"> Infiniti@ekimetrics.com </a> <o:p></o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'>Best regards,</p>

    <p class=MsoNormal style='text-align:justify'><b style='mso-bidi-font-weight:
    normal'>Infiniti Europe Aftersales team<o:p></o:p></b></p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-GB
    style='font-size:8.0pt;mso-bidi-font-size:11.0pt;line-height:107%;mso-ansi-language:
    EN-GB'>*Compliance ratio = the number of eligible vehicles with Printed<b
    style='mso-bidi-font-weight:normal'> Warranty certificate in system </b>/ <b
    style='mso-bidi-font-weight:normal'>Total number of eligible vehicles<o:p></o:p></b></span></p>

    </div>

    </body>

    </html>"""

        elif self.country == "France":

            if self.compliance_ratio == 100:

    ### TEXTE FRANÇAIS 100%

                self.text = """<html>

    <head>

    </head>

    <body lang=FR style='tab-interval:36.0pt'>

    <div class=WordSection1>

    <p class=MsoNormal><span lang=EN-US>Cher Centre Infiniti""" + self.center + """</span></p>

    <p class=MsoNormal><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>Nous voici arrivés à la fin du mois de
     """+ self.month + """ et nous aimerions vous partager votre 
    <b style='mso-bidi-font-weight:normal'>taux de conformité* par rapport au certificat de garantie imprimé
    </b> sur l’ensemble des véhicules éligibles.</span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>Comme vous le savez, <b
    style='mso-bidi-font-weight:normal'><u>l'objectif est d'atteindre un taux de conformité de 100 %.<o:p></o:p></u></b></span></p>

    <p class=MsoNormal align=center style='text-align:center'><b style='mso-bidi-font-weight:
    normal'><span lang=EN-US><o:p>&nbsp;</o:p></span></b></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>À la fin du mois, vous avez atteint un taux de conformité de : 
    <b style='mso-bidi-font-weight:normal'> 100% <o:p></o:p></u></b></span></p>

    <p class=MsoNormal><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>Merci de remplir les exigences de garantie de prévente pour les véhicules vendus. 
    Vous avez réalisé l'objectif de ce mois-ci. Poursuivez vos efforts afin de continuer à offrir un service de qualité.</span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>Sincères salutations,</span></p>

    <p class=MsoNormal style='text-align:justify'><b style='mso-bidi-font-weight:
    normal'>Infiniti Europe Aftersales team<o:p></o:p></b></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-GB
    style='font-size:8.0pt;mso-bidi-font-size:11.0pt;line-height:107%;mso-ansi-language:
    EN-GB'>*Taux de conformité = Quantité de véhicules éligibles possédant un certificat de garantie imprimé dans 
    le système / Total des véhicules éligibles<o:p></o:p></span></p>

    </div>

    </body>

    </html>"""

            else:

    ### TEXTE FRANÇAIS NOT 100%

                self.text = """<html>

    <head>

    </head>

    <body lang=EN-US link="#0563C1" vlink="#954F72" style='tab-interval:36.0pt'>

    <div class=WordSection1>

    <p class=MsoNormal>Cher Centre Infiniti """ + self.center + """,</p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'>Comme vous le savez, tous les véhicules vendus doivent remplir les exigences 
    de garantie de prévente, et leur certificat de garantie doit être imprimé et annexé au Manuel d’utilisation. Le certificat 
    de garantie ne peut être imprimé que si toutes les exigences de garantie (WB-IEU-17-002) sont remplies.</p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'>Votre <b style='mso-bidi-font-weight:
    normal'>taux de conformité* actuel pour les certificats de garantie imprimés</b> sur l’ensemble des véhicules éligibles est de :
    <b style='mso-bidi-font-weight:normal'> """ + str(self.compliance_ratio) + """%</b></p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'><b style='mso-bidi-font-weight:
    normal'><u>L'objectif est d'atteindre un taux de conformité de 100 %.<o:p></o:p></u></b></p>

    <p class=MsoNormal style='text-align:justify'>Vous trouverez en annexe une liste de VIN qui ont déjà été livrés aux 
    clients, mais dont la date de début de garantie n'a pas été saisie dans le système et/ou dont le certificat de garantie n'a pas été imprimé. 
    Nous avons divisé le rapport en deux sections.</p>

    <p class=MsoListParagraphCxSpFirst style='margin-left:38.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-18.0pt;mso-list:l1 level1 lfo5'><![if !supportLists]><span
    style='font-family:Symbol;mso-fareast-font-family:Symbol;mso-bidi-font-family:
    Symbol'><span style='mso-list:Ignore'>·<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>•    Section 1 : Les VIN dont la <b
    style='mso-bidi-font-weight:normal'>date de début de garantie </b>est manquante et dont le
     <b style='mso-bidi-font-weight:normal'>certificat de garantie n'a pas été imprimé.</b></p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:74.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-18.0pt;mso-list:l1 level2 lfo5'><![if !supportLists]><span
    style='font-family:"Courier New";mso-fareast-font-family:"Courier New"'><span
    style='mso-list:Ignore'>o<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;
    </span></span></span><![endif]><b style='mso-bidi-font-weight:normal'>Actions requises : </b></p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:110.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-30.0pt;mso-list:l1 level3 lfo5'><![if !supportLists]><span
    style='mso-bidi-font-family:Calibri;mso-bidi-theme-font:minor-latin'><span
    style='mso-list:Ignore'>1.<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>Saisissez la date de début de garantie dans votre système de garantie Infiniti. Il est à noter 
    que la date de début de garantie est la date de remise du véhicule au client ou la date à laquelle le véhicule est mis en circulation.</p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:110.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-30.0pt;line-height:105%;mso-list:l1 level3 lfo5'><![if !supportLists]><span
    style='mso-bidi-font-family:Calibri;mso-bidi-theme-font:minor-latin'><span
    style='mso-list:Ignore'>2.<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>Imprimez le certificat de garantie et remettez-le au client lors de sa prochaine visite au Centre Infiniti.</p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:38.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-18.0pt;mso-list:l1 level1 lfo5'><![if !supportLists]><span
    style='font-family:Symbol;mso-fareast-font-family:Symbol;mso-bidi-font-family:
    Symbol'><span style='mso-list:Ignore'>·<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>•    Section 2 : Les VIN dont le <b
    style='mso-bidi-font-weight:normal'>certificat de garantie </b>doit être imprimé.</p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:74.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-18.0pt;mso-list:l1 level2 lfo5'><![if !supportLists]><span
    style='font-family:"Courier New";mso-fareast-font-family:"Courier New"'><span
    style='mso-list:Ignore'>o<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;
    </span></span></span><![endif]><b style='mso-bidi-font-weight:normal'>Actions requises : <o:p></o:p></b></p>

    <p class=MsoListParagraphCxSpLast style='margin-left:110.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-30.0pt;line-height:105%;mso-list:l1 level3 lfo5'><![if !supportLists]><span
    style='mso-bidi-font-family:Calibri;mso-bidi-theme-font:minor-latin'><span
    style='mso-list:Ignore'>1.<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>Imprimez le certificat de garantie et remettez-le au client lors de sa prochaine visite au Centre Infiniti.</p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'>Vous trouverez en annexe les exigences de garantie (PSWR), 
    avec tous les détails concernant les tâches à exécuter impérativement avant de remettre un véhicule 
    au client final. La couverture de garantie pourra alors être validée et le nouveau véhicule du client 
    sera également couvert par le programme d'assistance routière. Merci de veiller à satisfaire toutes les 
    exigences de garantie <b style='mso-bidi-font-weight:normal'><u>avant</u></b> de remettre le véhicule au client.</p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=3DMsoNormal><span class=3DGramE><span lang=3DEN-US style=3D'mso-an=
    si-language:
    EN-US'>Si</span></span><span lang=3DEN-US style=3D'mso-ansi-language:EN-US'=
    > vous avez des questions n'hésitez pas à contacter : <a href="mailto:Infiniti@ekimetrics.com"> Infiniti@ekimetrics.com </a> <o:p></o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'>Sincères salutations,</p>

    <p class=MsoNormal style='text-align:justify'><b style='mso-bidi-font-weight:
    normal'>Infiniti Europe Aftersales team<o:p></o:p></b></p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-GB
    style='font-size:8.0pt;mso-bidi-font-size:11.0pt;line-height:107%;mso-ansi-language:
    EN-GB'>*Taux de conformité = Quantité de véhicules éligibles possédant un certificat de garantie<b
    style='mso-bidi-font-weight:normal'>  imprimé rempli dans le système </b>/ <b
    style='mso-bidi-font-weight:normal'>Total des véhicules éligibles<o:p></o:p></b></span></p>

    </div>

    </body>

    </html>"""

        elif self.country == "Italy":

            if self.compliance_ratio == 100:

    ### TEXTE ITALIEN 100%

                self.text = """<html>

    <head>

    </head>

    <body lang=FR style='tab-interval:36.0pt'>

    <div class=WordSection1>

    <p class=MsoNormal><span lang=EN-US>Gentile Centro Infiniti """ + self.center + """</span></p>

    <p class=MsoNormal><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>Siamo alla fine del mese di 
    """+ self.month + """ e vorremmo condividere con Lei un aggiornamento sul Suo 
    <b style='mso-bidi-font-weight:normal'>tasso di conformità* relativo al certificato di garanzia stampato 
    </b> per tutti i veicoli idonei.</span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>Come saprà, <b
    style='mso-bidi-font-weight:normal'><u>il tasso di conformità ha come obiettivo il raggiungimento del 100%.<o:p></o:p></u></b></span></p>

    <p class=MsoNormal align=center style='text-align:center'><b style='mso-bidi-font-weight:
    normal'><span lang=EN-US><o:p>&nbsp;</o:p></span></b></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>Alla fine del mese il Suo tasso di conformità risulta essere del: 
    <b style='mso-bidi-font-weight:normal'> 100% <o:p></o:p></u></b></span></p>

    <p class=MsoNormal><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>La ringraziamo per aver rispettato i requisiti di garanzia prevendita dei veicoli 
    venduti / consegnati al cliente. Ha raggiunto l’obiettivo di questo mese. Continui a impegnarsi per garantire sempre l’alta qualità del servizio.</span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US>Cordiali saluti,</span></p>

    <p class=MsoNormal style='text-align:justify'><b style='mso-bidi-font-weight:
    normal'>Infiniti Europe Aftersales team<o:p></o:p></b></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-GB
    style='font-size:8.0pt;mso-bidi-font-size:11.0pt;line-height:107%;mso-ansi-language:
    EN-GB'>* Tasso di conformità = Numero di veicoli idonei con certificato di garanzia stampato a sistema / Totale dei veicoli idonei<o:p></o:p></span></p>

    </div>

    </body>

    </html>"""

            else:

    ### TEXTE ITALIEN NOT 100%

                self.text = """<html>

    <head>

    </head>

    <body lang=EN-US link="#0563C1" vlink="#954F72" style='tab-interval:36.0pt'>

    <div class=WordSection1>

    <p class=MsoNormal>Gentile Centro Infiniti """ + self.center + """,</p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'>Come sapete, i requisiti di garanzia prevendita 
    devono essere completi per tutti i veicoli venduti / consegnati ai clienti e il certificato di garanzia deve essere 
    stampato e allegato al relativo libretto. Il certificato di garanzia non può essere stampato senza che i requisiti di garanzia 
    prevendita siano completi (WB-IEU-17-002).</p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'>Il vostro attuale <b style='mso-bidi-font-weight:
    normal'>tasso di conformità* relativo al certificato di garanzia stampato</b> per tutti i veicoli idonei è del:
    <b style='mso-bidi-font-weight:normal'> """ + str(self.compliance_ratio) + """%</b></p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'><b style='mso-bidi-font-weight:
    normal'><u>Il tasso di conformità ha come obiettivo il raggiungimento del 100%.<o:p></o:p></u></b></p>

    <p class=MsoNormal style='text-align:justify'>In allegato potete trovare un elenco di VIN già consegnati al cliente, ma ancora 
    privi di una data di garanzia a sistema e/o di un certificato di garanzia stampato. Abbiamo suddiviso il report in due sezioni.</p>

    <p class=MsoListParagraphCxSpFirst style='margin-left:38.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-18.0pt;mso-list:l1 level1 lfo5'><![if !supportLists]><span
    style='font-family:Symbol;mso-fareast-font-family:Symbol;mso-bidi-font-family:
    Symbol'><span style='mso-list:Ignore'>·<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>•    Sezione 1: VIN senza <b
    style='mso-bidi-font-weight:normal'>data di inizio della garanzia </b>e senza 
     <b style='mso-bidi-font-weight:normal'>certificato di garanzia stampato.</b></p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:74.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-18.0pt;mso-list:l1 level2 lfo5'><![if !supportLists]><span
    style='font-family:"Courier New";mso-fareast-font-family:"Courier New"'><span
    style='mso-list:Ignore'>o<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;
    </span></span></span><![endif]><b style='mso-bidi-font-weight:normal'>Interventi richiesti : </b></p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:110.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-30.0pt;mso-list:l1 level3 lfo5'><![if !supportLists]><span
    style='mso-bidi-font-family:Calibri;mso-bidi-theme-font:minor-latin'><span
    style='mso-list:Ignore'>1.<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>Inserire la data di inizio della garanzia nel sistema garanzie Infiniti. Si prega di notare che 
    la data di inizio della garanzia coincide con la data di consegna al cliente o con la data di primo utilizzo del veicolo, a seconda 
    di quale condizione si verifica per prima.</p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:110.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-30.0pt;line-height:105%;mso-list:l1 level3 lfo5'><![if !supportLists]><span
    style='mso-bidi-font-family:Calibri;mso-bidi-theme-font:minor-latin'><span
    style='mso-list:Ignore'>2.<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>Stampare il certificato di garanzia. Fornire quindi il certificato di garanzia al 
    cliente alla prima occasione in cui questi fa visita al vostro Centro Infiniti.</p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:38.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-18.0pt;mso-list:l1 level1 lfo5'><![if !supportLists]><span
    style='font-family:Symbol;mso-fareast-font-family:Symbol;mso-bidi-font-family:
    Symbol'><span style='mso-list:Ignore'>·<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>•    Sezione 2: VIN per i quali si deve stampare il <b
    style='mso-bidi-font-weight:normal'>certificato di garanzia </b></p>

    <p class=MsoListParagraphCxSpMiddle style='margin-left:74.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-18.0pt;mso-list:l1 level2 lfo5'><![if !supportLists]><span
    style='font-family:"Courier New";mso-fareast-font-family:"Courier New"'><span
    style='mso-list:Ignore'>o<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;
    </span></span></span><![endif]><b style='mso-bidi-font-weight:normal'>Interventi richiesti : <o:p></o:p></b></p>

    <p class=MsoListParagraphCxSpLast style='margin-left:110.35pt;mso-add-space:
    auto;text-align:justify;text-indent:-30.0pt;line-height:105%;mso-list:l1 level3 lfo5'><![if !supportLists]><span
    style='mso-bidi-font-family:Calibri;mso-bidi-theme-font:minor-latin'><span
    style='mso-list:Ignore'>1.<span style='font:7.0pt "Times New Roman"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    </span></span></span><![endif]>Stampare il certificato di garanzia. Fornire quindi il certificato di garanzia al cliente 
    alla prima occasione in cui questi fa visita al vostro Centro Infiniti.</p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-US><o:p>&nbsp;</o:p></span></p>

    <p class=MsoNormal style='text-align:justify'>In allegato potete trovare i requisiti di garanzia prevendita (PSWR) con tutti i 
    dettagli relativi alle operazioni obbligatorie da completare prima di consegnare un veicolo al cliente finale, in modo da validare 
    la copertura della garanzia e dotare il nuovo veicolo del cliente della copertura di assistenza stradale. Assicurarsi che tutti i 
    requisiti di garanzia siano completi <b style='mso-bidi-font-weight:normal'><u>prima</u></b> della consegna al cliente.</p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=3DMsoNormal><span class=3DGramE><span lang=3DEN-US style=3D'mso-an=
    si-language:
    EN-US'>Qualora</span></span><span lang=3DEN-US style=3D'mso-ansi-language:EN-US'=
    >abbia domande, non esiti a contattar : <a href="mailto:Infiniti@ekimetrics.com"> Infiniti@ekimetrics.com </a> <o:p></o:p></span></p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'>Cordiali saluti,</p>

    <p class=MsoNormal style='text-align:justify'><b style='mso-bidi-font-weight:
    normal'>Infiniti Europe Aftersales team<o:p></o:p></b></p>

    <p class=MsoNormal style='text-align:justify'><o:p>&nbsp;</o:p></p>

    <p class=MsoNormal style='text-align:justify'><span lang=EN-GB
    style='font-size:8.0pt;mso-bidi-font-size:11.0pt;line-height:107%;mso-ansi-language:
    EN-GB'>*Tasso di conformità = Numero di veicoli idonei con certificato di garanzia <b
    style='mso-bidi-font-weight:normal'>stampato e compilato a sistema </b>/ <b
    style='mso-bidi-font-weight:normal'>Totale dei veicoli idonei<o:p></o:p></b></span></p>

    </div>

    </body>

    </html>"""










