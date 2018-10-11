#!/usr/bin/env python
# -*- coding: utf-8 -*- 


"""--------------------------------------------------------------------
INFINITI WARRANTY PROJECT


Hugo PERRIN
------------------------------------------------------------------------
"""


# ### LIBRARIES

# Usual Libraries
import sys
sys.path.append("S:/121. Infiniti/Scripts/Full_Automatique")

from datetime import date, datetime
import pandas as pd
from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import numpy as np
from email import *
import win32com.client as win32
import time

# Custom libraries
from Infiniti import *
time1 = time.time()




#=============================================================================================================================
# MAIN LOOP
#=============================================================================================================================



if __name__ == '__main__':

    #=============================================================================================================================
    # MAIN VARIABLES
    #============================================================================================================================= 

    # YEAR/MONTH
    date = date.today()
    year = str(date.year)
    month = str(date.month)    

    # DATA PATH
    path_main = 'S:/121. Infiniti/Excel'
    path_bilan = 'S:/121. Infiniti/Bilan'
    path_attachments = 'S:/121. Infiniti/Fichiers à envoyer aux dealers'
    path_bulletin = 'S:/121. Infiniti/PDF'
    path_contact = 'S:/121. Infiniti/Contacts'

    path_main = os.path.join(path_main, year, month)
    path_bilan = os.path.join(path_bilan, year, month)
    path_attachments = os.path.join(path_attachments, year, month)

    # DATA LOADING : handle sources manually !!!!!!!!!!!

    # Global data
    print(">> Global Registration file loading",end = "...")

    global_registration = pd.read_excel(path_main + "/Infiniti - Sales vs PSWR Performance Jan 16 - Aug 17.xlsx", "Status on 1 September 2017")

    print("done")
    print(">> France data file loading", end = "...")

    France_base = pd.read_excel(path_main + "/France_VINS.xlsx")
    France_vin_name = "VIN"

    print("done")
    print(">> Italy data file loading", end = "...")

    Italy_base1 = pd.read_excel(path_main + "/Main.xlsx", "IT CY2016")
    Italy_base2 = pd.read_excel(path_main + "/Main.xlsx", "IT FY2016")
    Italy_base3 = pd.read_excel(path_main + "/Italy_06092017.xlsx", "IT")
    Italy_base = pd.concat([Italy_base1,Italy_base2, Italy_base3], axis = 0)
    Italy_vin_name = "VIN"

    print("done")
    print(">> UK data file loading", end = "...")

    UK_base1 = pd.read_excel(path_main + "/Main.xlsx", "UK 2015")
    UK_base2 = pd.read_excel(path_main + "/Main.xlsx", "UK 2016")
    UK_base3 = pd.read_excel(path_main + "/UK Vins.xlsx") 
    UK_base = pd.concat([UK_base1,UK_base2,UK_base3], axis = 0)
    UK_vin_name = "Vin Chasis Frame Number"

    print("done")
    print(">> Contact database loading", end = "...")

    ## Test pour l'envoie des mails
    # contact_sheet = pd.read_excel(path_contact + "/contact_database.xlsx", "perso") 
    # contact_sheet = pd.read_excel(path_contact + "/contact_database.xlsx", "strasbourg")        
    ## !!!!!!!   Test officiel    !!!!!!!!
    # contact_sheet = pd.read_excel(path_contact + "/contact_database.xlsx", "infiniti")       
    ## !!!!!!!!!!!!!!!!!!   ENVOI OFFICIEL    !!!!!!!!!!!!!!
    contact_sheet = pd.read_excel(path_contact + "/contact_database.xlsx", "dealer")   

    print("done\n")


    # On règle quelques détails : 

    #  PROBLEME DE DUPLICATION AVEC DIJON
    global_registration["Dealer"] = global_registration["Dealer"].apply(lambda x: "INFINITI CENTER DIJON" if x == "INFINITI CENTRE DIJON" else x)


    #=============================================================================================================================
    # DEALER SHEET TEMPLATE
    #============================================================================================================================= 


    # On crée un workbook "bilan"
    bilan = Workbook()
    dealer_sheet = bilan.active
    dealer_sheet.title = "Dealer"

    # Mise en forme de l'entête du tableau principal
    dealer_sheet.cell(row = 2, column = 2, value = "RESULTS PER DEALER")

    style_range(dealer_sheet, row_min = 2, row_max = 2, col_min = 2, col_max = 2,
                fill = PatternFill(fill_type='solid', start_color='043C65'),
                font = Font(name='Calibri', size=10, bold=False, color = 'FFFFFF'),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

    dealer_sheet.cell(row = 4, column = 2, value = "Country")
    dealer_sheet.cell(row = 4, column = 3, value = "Dealer")
    dealer_sheet.cell(row = 4, column = 4, value = "Number of eligible/register vehicles")
    dealer_sheet.cell(row = 4, column = 5, value = "Number of Vins not registered")
    dealer_sheet.cell(row = 4, column = 6, value = "Number of eligible vehicles with a warranty start date")
    dealer_sheet.cell(row = 4, column = 7, value = "Number of eligible vehicles with a warranty certificate")
    dealer_sheet.cell(row = 4, column = 8, value = "Compliance Ratio")

    style_range(dealer_sheet, row_min = 4, row_max = 4,col_min = 2, col_max = 8,
                fill = PatternFill(fill_type='solid', start_color='043C65'),
                font = Font(name='Calibri', size=10, bold=False, color = 'FFFFFF'),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

    set_columns_width(dealer_sheet,20,[1,3,5,7])
    set_columns_width(dealer_sheet,30,[2,4,6])
    set_columns_width(dealer_sheet,5,[0])



    #=============================================================================================================================
    # DEALERS RESULTS
    #============================================================================================================================= 



    # Process pour chaque dealer de global registration: 
    # 
    # 1 - On obtient le pays :
    #     - Si ce n'est pas France, Italy ou UK on s'arrête.
    # 2 - On obtient le compliance ratio : 
    #     - Si c'est 100% on s'arrête.
    # 3 - On crée un workbook "Dealer".
    # 
    # 4 - On obtient les WSD, les WC et les RNS.
    # 
    # 5 - On met les WSD et les WC dans deux feuilles séparées du workbook "Dealer".
    # 
    # 6 - On met dans le workbook "bilan" les quantités correspondantes.
    # 
    # 7 - On sauvegarde la pièce jointe "DealerName.xlsx".
    # 
    # 8 - On envoie le mail à l'adresse correspondante.
    # 
    # Une fois la boucle terminée, on envoie le bilan !!!

    # Note : LE BILAN DOIT ETRE ENVOYE AVANT LES MAILS AUTOMATIQUES 


    #On initialise les compteurs généraux

    eligible_number_fr = 0
    RNS_number_fr = 0
    WSD_number_fr = 0
    WC_number_fr = 0

    eligible_number_UK = 0
    RNS_number_UK = 0
    WSD_number_UK = 0
    WC_number_UK = 0

    eligible_number_It = 0
    RNS_number_It = 0
    WSD_number_It = 0
    WC_number_It = 0
            
    eligible_number = 0
    RNS_number = 0
    WSD_number = 0
    WC_number = 0

    dealer = {}


    # Boucle sur les dealers : tout est sauvegardé dans un dictionnaire pour pouvoir être envoyé 
    # plus tard par mail automatique sans avoir à être recalculé 


    for i,name in enumerate(list(contact_sheet["Dealer"])):

        print(">> {} in treatment".format(name), end = "...")
        
        dealer[name] = Dealer(name, global_registration, path_attachments, contact_sheet)
        
        if dealer[name].country == "France":
            dealer[name].get_RNSvsEligible(France_base, France_vin_name)
        elif dealer[name].country == "UK":
            dealer[name].get_RNSvsEligible(UK_base, UK_vin_name)
        elif dealer[name].country == "Italy":
            dealer[name].get_RNSvsEligible(Italy_base, Italy_vin_name)
            
        dealer[name].get_lacking_WSD()
        
        dealer[name].get_lacking_WC()
        
        dealer[name].get_compliance_ratio()
        
        dealer[name].get_contact_mail(contact_sheet)

        if dealer[name].contact is not None:

            dealer[name].create_attachment()
        
            dealer[name].save_attachment()

        # Remplissage du tableau principal
        dealer_sheet.cell(row = 5+i, column = 2, value = dealer[name].country)
        dealer_sheet.cell(row = 5+i, column = 3, value = dealer[name].name)
        dealer_sheet.cell(row = 5+i, column = 4, value = dealer[name].eligible_number)
        dealer_sheet.cell(row = 5+i, column = 5, value = dealer[name].RNS_number)
        dealer_sheet.cell(row = 5+i, column = 6, value = dealer[name].WSD_number)
        dealer_sheet.cell(row = 5+i, column = 7, value = dealer[name].WC_number)
        dealer_sheet.cell(row = 5+i, column = 8, value = dealer[name].compliance_ratio/100)
        
        
        # Calcul des valeurs du tableau secondaire
        if dealer[name].country == "France":
            eligible_number_fr += dealer[name].eligible_number
            RNS_number_fr += dealer[name].RNS_number
            WSD_number_fr += dealer[name].WSD_number
            WC_number_fr += dealer[name].WC_number
        elif dealer[name].country == "UK":
            eligible_number_UK += dealer[name].eligible_number
            RNS_number_UK += dealer[name].RNS_number
            WSD_number_UK += dealer[name].WSD_number
            WC_number_UK += dealer[name].WC_number
        elif dealer[name].country == "Italy":
            eligible_number_It += dealer[name].eligible_number
            RNS_number_It += dealer[name].RNS_number
            WSD_number_It += dealer[name].WSD_number
            WC_number_It += dealer[name].WC_number
            
        eligible_number += dealer[name].eligible_number
        RNS_number += dealer[name].RNS_number
        WSD_number += dealer[name].WSD_number
        WC_number += dealer[name].WC_number
        
        print("done")
            
    #=============================================================================================================================
    # BILAN SHEET TEMPLATE
    #=============================================================================================================================

    #=============================================================================================================================
    # GLOBAL RESULTS

    global_sheet = bilan.create_sheet("Global")

    # france
    global_sheet.cell(row = 5, column = 3, value = eligible_number_fr)
    global_sheet.cell(row = 5, column = 4, value = RNS_number_fr)
    global_sheet.cell(row = 5, column = 5, value = WSD_number_fr)    
    global_sheet.cell(row = 5, column = 6, value = WC_number_fr)
    global_sheet.cell(row = 5, column = 7, value = WC_number_fr/eligible_number_fr)

    # UK
    global_sheet.cell(row = 7, column = 3, value = eligible_number_UK)
    global_sheet.cell(row = 7, column = 4, value = RNS_number_UK)
    global_sheet.cell(row = 7, column = 5, value = WSD_number_UK)    
    global_sheet.cell(row = 7, column = 6, value = WC_number_UK)
    global_sheet.cell(row = 7, column = 7, value = WC_number_UK/eligible_number_UK)

    # Italy
    global_sheet.cell(row = 6, column = 3, value = eligible_number_It)
    global_sheet.cell(row = 6, column = 4, value = RNS_number_It)
    global_sheet.cell(row = 6, column = 5, value = WSD_number_It)    
    global_sheet.cell(row = 6, column = 6, value = WC_number_It)
    global_sheet.cell(row = 6, column = 7, value = WC_number_It/eligible_number_It)

    # Total
    global_sheet.cell(row = 9, column = 3, value = eligible_number)
    global_sheet.cell(row = 9, column = 4, value = RNS_number)
    global_sheet.cell(row = 9, column = 5, value = WSD_number)    
    global_sheet.cell(row = 9, column = 6, value = WC_number)
    global_sheet.cell(row = 9, column = 7, value = WC_number/eligible_number)

    print(" \n\n >> ALL DONE")


    # Mise en forme du tableau principal
    style_range(dealer_sheet, row_min = 5, row_max = 4+len(list(contact_sheet["Dealer"])),col_min = 4, col_max = 8,
                fill = PatternFill(fill_type='none', start_color='77B5FE'),
                font = Font(name='Calibri', size=10, bold=False),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False))

    style_range(dealer_sheet, row_min = 5, row_max = 4+len(list(contact_sheet["Dealer"])),col_min = 2, col_max = 3,
                fill = PatternFill(fill_type='solid', start_color='043C65'),
                font = Font(name='Calibri', size=10, bold=False, color = 'FFFFFF'),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=False))



    # ### RESULTATS GLOBAUX


    # Mise en forme du tableau secondaire
    global_sheet.cell(row = 2, column = 2, value = "GLOBAL RESULTS")

    style_range(global_sheet, row_min = 2, row_max = 2, col_min = 2, col_max = 2,
                fill = PatternFill(fill_type='solid', start_color='043C65'),
                font = Font(name='Calibri', size=10, bold=False, color = 'FFFFFF'),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False))

    global_sheet.cell(row = 4, column = 2, value = "Country")
    global_sheet.cell(row = 4, column = 3, value = "Number of eligible/register vehicles")
    global_sheet.cell(row = 4, column = 4, value = "Number of Vins not registered")
    global_sheet.cell(row = 4, column = 5, value = "Number of eligible vehicles with a warranty start date")
    global_sheet.cell(row = 4, column = 6, value = "Number of eligible vehicles with a warranty certificate")
    global_sheet.cell(row = 4, column = 7, value = "Compliance Ratio")

    style_range(global_sheet, row_min = 4, row_max = 4, col_min = 2, col_max = 7,
                fill = PatternFill(fill_type='solid', start_color='043C65'),
                font = Font(name='Calibri', size=10, bold=False, color = 'FFFFFF'),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

    global_sheet.cell(row = 5, column = 2, value = "France")
    global_sheet.cell(row = 6, column = 2, value = "Italy")
    global_sheet.cell(row = 7, column = 2, value = "UK")

    style_range(global_sheet, row_min = 5, row_max = 7, col_min = 2, col_max = 2,
                fill = PatternFill(fill_type='solid', start_color='043C65'),
                font = Font(name='Calibri', size=10, bold=False, color = 'FFFFFF'),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

    style_range(global_sheet, row_min = 5, row_max = 7,col_min = 3, col_max = 7,
                fill = PatternFill(fill_type='none', start_color='77B5FE'),
                font = Font(name='Calibri', size=10, bold=False),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

    global_sheet.cell(row = 9, column = 2, value = "TOTAL")

    style_range(global_sheet, row_min = 9, row_max = 9,col_min = 2, col_max = 2,
                fill = PatternFill(fill_type='solid', start_color='043C65'),
                font = Font(name='Calibri', size=10, bold=False, color = 'FFFFFF'),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

    style_range(global_sheet, row_min = 9, row_max = 9,col_min = 3, col_max = 7,
                fill = PatternFill(fill_type='none', start_color='77B5FE'),
                font = Font(name='Calibri', size=10, bold=False),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

    set_columns_width(global_sheet, 30,[5])
    set_columns_width(global_sheet, 20,[1,2,3,4,6])
    set_columns_width(global_sheet, 5,[0])

    #=============================================================================================================================
    # RETAIL GROUP RESULTS

    retail_sheet = bilan.create_sheet("Retail Group")

    dealer_concerned = ["INFINITI CENTRE GLASGOW",
                        "INFINITI CENTRE STOCKPORT",
                        "INFINITI CENTRE READING",
                        "INFINITI CENTRE LEEDS",
                        "INFINITI CENTRE BIRMINGHAM"]

    retail_sheet.cell(row = 2, column = 2, value = "RETAIL GROUP RESULTS")

    style_range(retail_sheet, row_min = 2, row_max = 2, col_min = 2, col_max = 2,
                fill = PatternFill(fill_type='solid', start_color='043C65'),
                font = Font(name='Calibri', size=10, bold=False, color = 'FFFFFF'),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False))

    retail_sheet.cell(row = 4, column = 2, value = "Dealer")
    retail_sheet.cell(row = 4, column = 3, value = "Number of eligible/register vehicles")
    retail_sheet.cell(row = 4, column = 4, value = "Number of Vins not registered")
    retail_sheet.cell(row = 4, column = 5, value = "Number of eligible vehicles with a warranty start date")
    retail_sheet.cell(row = 4, column = 6, value = "Number of eligible vehicles with a warranty certificate")
    retail_sheet.cell(row = 4, column = 7, value = "Compliance Ratio")

    style_range(retail_sheet, row_min = 4, row_max = 4, col_min = 2, col_max = 7,
                fill = PatternFill(fill_type='solid', start_color='043C65'),
                font = Font(name='Calibri', size=10, bold=False, color = 'FFFFFF'),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))                    

    eligible_number = 0
    RNS_number = 0
    WSD_number = 0
    WC_number = 0

    for i,dealer_retail in enumerate(dealer_concerned): 

        retail_sheet.cell(row = 5+i, column = 2, value = dealer[dealer_retail].name)
        retail_sheet.cell(row = 5+i, column = 3, value = dealer[dealer_retail].eligible_number)
        retail_sheet.cell(row = 5+i, column = 4, value = dealer[dealer_retail].RNS_number)
        retail_sheet.cell(row = 5+i, column = 5, value = dealer[dealer_retail].WSD_number)    
        retail_sheet.cell(row = 5+i, column = 6, value = dealer[dealer_retail].WC_number)
        retail_sheet.cell(row = 5+i, column = 7, value = dealer[dealer_retail].compliance_ratio/100)

        eligible_number += dealer[dealer_retail].eligible_number
        RNS_number += dealer[dealer_retail].RNS_number
        WSD_number += dealer[dealer_retail].WSD_number
        WC_number += dealer[dealer_retail].WC_number

    compliance_ratio = WC_number/eligible_number

    style_range(retail_sheet, row_min = 5, row_max = 9, col_min = 2, col_max = 2,
                fill = PatternFill(fill_type='solid', start_color='043C65'),
                font = Font(name='Calibri', size=10, bold=False, color = 'FFFFFF'),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

    style_range(retail_sheet, row_min = 5, row_max = 9,col_min = 3, col_max = 7,
                fill = PatternFill(fill_type='none', start_color='77B5FE'),
                font = Font(name='Calibri', size=10, bold=False),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

    retail_sheet.cell(row = 11, column = 2, value = "TOTAL")

    retail_sheet.cell(row = 11, column = 3, value = eligible_number)
    retail_sheet.cell(row = 11, column = 4, value = RNS_number)
    retail_sheet.cell(row = 11, column = 5, value = WSD_number)    
    retail_sheet.cell(row = 11, column = 6, value = WC_number)
    retail_sheet.cell(row = 11, column = 7, value = compliance_ratio)


    style_range(retail_sheet, row_min = 11, row_max = 11,col_min = 2, col_max = 2,
                fill = PatternFill(fill_type='solid', start_color='043C65'),
                font = Font(name='Calibri', size=10, bold=False, color = 'FFFFFF'),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))

    style_range(retail_sheet, row_min = 11, row_max = 11,col_min = 3, col_max = 7,
                fill = PatternFill(fill_type='none', start_color='77B5FE'),
                font = Font(name='Calibri', size=10, bold=False),
                border = Border(left=Side(border_style="thin", color='FF000000'),
                                right=Side(border_style="thin", color='FF000000'),
                                top=Side(border_style="thin", color='FF000000'),
                                bottom=Side(border_style="thin", color='FF000000')),
                alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True))


    set_columns_width(retail_sheet, 30,[1,5])
    set_columns_width(retail_sheet, 20,[2,3,4,6])
    set_columns_width(retail_sheet, 5,[0])


    #Sauvegarde
    bilan.save(filename = path_bilan + "/Bilan.xlsx")


    #=============================================================================================================================
    # RETAIL GROUP RESULTS













    # Attention, l'envoie du bilan n'est pas automatique, il y a un travail de mise en forme à faire au préalable : 
    #    - Figer les volets.
    #    - Mise en forme conditionnelle des compliances ratio.
    #    - Tri par pays puis par compliance ratio.
    #    - Ecrire en "metric bold" pour les titres puis en "metric regular" pour le reste.
    #    - Mettre un séparateur de millier pour les nombres.

    #### COPIER LA MISE EN FORME DES ANCIENS BILAN

    # Temps total
    time2 = time.time()
    diff_time = (time2 - time1)/60
    print("Le temps d'exécution du programme est de {} minutes\n\n".format(round(diff_time,1)))


    # Note : Ne pas fermer le programme avant l'envoie des mails, sinon il faudra tout relancer 

    print("Careful to whom you are sending the emails and check the results before sending them \n")
    # launch = input(">> Do you want to send the emails ? [Yes/No]\n")
    launch = "Yes"


    # ### Automatic emailing of Infiniti network
    if launch == "Yes":

        time1 = time.time()

        for name in list(contact_sheet["Dealer"]):
            
            try:
                dealer[name].send_email(contact_sheet)
            except:
                pass
            
        time2 = time.time()
        diff_time = (time2 - time1)/60
        print("Le temps d'envoie des mails automatiques est de {} minutes".format(round(diff_time,1)))





